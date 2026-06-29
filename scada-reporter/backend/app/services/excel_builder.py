from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill

from app.i18n import get_labels

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FLOAT_FMT = "#,##0.000"


def _header_row(ws, cols: list[str], row: int = 1):
    for col_idx, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _embed_image(ws, png_bytes: bytes, anchor: str, width_px: int = 700, height_px: int = 245):
    img = XLImage(BytesIO(png_bytes))
    img.width = width_px
    img.height = height_px
    ws.add_image(img, anchor)


def build_advanced_excel(
    archive,
    per_tag_data: list[dict],
    template,
    summary_chart_png: bytes,
    lang: str = "en",
    grafana_charts: list[dict] | None = None,
    variables: list[dict] | None = None,
) -> bytes:
    L = get_labels(lang)  # noqa: N806 — short alias for the label dict, used pervasively below
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_ozet = wb.active
    ws_ozet.title = L["summary_sheet"]

    if template.show_summary_stats:
        stat_headers = [
            L["tag"],
            L["unit"],
            L["count"],
            L["good_quality"],
            L["availability_pct"],
            L["average"],
            L["std_dev"],
            L["minimum"],
            L["maximum"],
            L["trend"],
            L["trend_r2"],
            L["anomaly_count"],
        ]
        _header_row(ws_ozet, stat_headers, row=1)
        for row_idx, td in enumerate(per_tag_data, start=2):
            tag = td["tag"]
            s = td["stats"]
            ws_ozet.cell(row=row_idx, column=1, value=tag.name)
            ws_ozet.cell(row=row_idx, column=2, value=tag.unit or "")
            ws_ozet.cell(row=row_idx, column=3, value=s.count)
            ws_ozet.cell(row=row_idx, column=4, value=s.good_quality_count)
            c = ws_ozet.cell(row=row_idx, column=5, value=round(s.availability_pct, 2))
            c.number_format = FLOAT_FMT
            for col_off, val in enumerate([s.avg, s.std_dev, s.min, s.max], start=6):
                c = ws_ozet.cell(
                    row=row_idx, column=col_off, value=round(val, 3) if val is not None else None
                )
                c.number_format = FLOAT_FMT
            ws_ozet.cell(row=row_idx, column=10, value=s.trend_direction)
            c = ws_ozet.cell(
                row=row_idx,
                column=11,
                value=round(s.trend_r2, 4) if s.trend_r2 is not None else None,
            )
            c.number_format = FLOAT_FMT
            ws_ozet.cell(row=row_idx, column=12, value=len(td["anomalies"]))

        # embed summary bar chart below table
        chart_row = len(per_tag_data) + 4
        if summary_chart_png:
            _embed_image(ws_ozet, summary_chart_png, f"A{chart_row}")

    # --- Per-tag sheets ---
    for td in per_tag_data:
        tag = td["tag"]
        s = td["stats"]
        anomalies = td["anomalies"]
        period_rows = td["period_rows"]
        chart_png = td.get("chart_png", b"")

        sheet_name = tag.name[:31]
        ws = wb.create_sheet(title=sheet_name)
        current_row = 1

        # Stats block
        ws.cell(row=current_row, column=1, value=L["statistics"]).font = HEADER_FONT
        current_row += 1
        for label, val in [
            (L["tag"], tag.name),
            (L["unit"], tag.unit or ""),
            (L["total_reads"], s.count),
            (L["good_quality"], s.good_quality_count),
            (
                L["availability_pct"],
                round(s.availability_pct, 2) if s.availability_pct is not None else None,
            ),
            (L["average"], round(s.avg, 3) if s.avg is not None else None),
            (L["std_dev"], round(s.std_dev, 3) if s.std_dev is not None else None),
            (L["minimum"], round(s.min, 3) if s.min is not None else None),
            (L["maximum"], round(s.max, 3) if s.max is not None else None),
            (L["trend"], s.trend_direction),
            (
                L["trend_slope"],
                round(s.trend_slope, 5) if s.trend_slope is not None else None,
            ),
            (L["trend_r2"], round(s.trend_r2, 4) if s.trend_r2 is not None else None),
            (L["gap_count"], s.gap_count),
            (L["gap_total_seconds"], round(s.gap_total_seconds, 1)),
        ]:
            ws.cell(row=current_row, column=1, value=label)
            c = ws.cell(row=current_row, column=2, value=val)
            if isinstance(val, float):
                c.number_format = FLOAT_FMT
            current_row += 1

        # Percentiles
        if s.percentiles:
            current_row += 1
            ws.cell(row=current_row, column=1, value=L["percentiles"]).font = HEADER_FONT
            current_row += 1
            for level, pval in sorted(s.percentiles.items()):
                ws.cell(row=current_row, column=1, value=f"P{level}")
                c = ws.cell(row=current_row, column=2, value=round(pval, 3))
                c.number_format = FLOAT_FMT
                current_row += 1

        # Anomaly sub-table
        if template.show_anomaly_table and anomalies:
            current_row += 1
            ws.cell(row=current_row, column=1, value=L["anomalies"]).font = HEADER_FONT
            current_row += 1
            _header_row(
                ws,
                [L["time"], L["value"], L["type"], L["severity"], L["detail"]],
                row=current_row,
            )
            current_row += 1
            for ev in anomalies:
                ws.cell(row=current_row, column=1, value=ev.timestamp.isoformat())
                c = ws.cell(
                    row=current_row,
                    column=2,
                    value=round(ev.value, 3) if ev.value is not None else None,
                )
                c.number_format = FLOAT_FMT
                ws.cell(row=current_row, column=3, value=ev.anomaly_type)
                ws.cell(row=current_row, column=4, value=ev.severity)
                ws.cell(row=current_row, column=5, value=ev.details)
                current_row += 1

        # Timeseries chart
        if template.show_trend_charts and chart_png:
            current_row += 1
            _embed_image(ws, chart_png, f"A{current_row}")
            current_row += 15  # reserve rows for image

        # Period aggregation table
        if period_rows:
            current_row += 1
            ws.cell(row=current_row, column=1, value=L["period_summary"]).font = HEADER_FONT
            current_row += 1
            _header_row(
                ws,
                [L["period"], L["average"], L["minimum"], L["maximum"], L["count"]],
                row=current_row,
            )
            current_row += 1
            for pr in period_rows:
                ws.cell(row=current_row, column=1, value=pr["period"])
                for col_off, key in enumerate(["mean", "min", "max"], start=2):
                    c = ws.cell(row=current_row, column=col_off, value=pr[key])
                    c.number_format = FLOAT_FMT
                ws.cell(row=current_row, column=5, value=pr["count"])
                current_row += 1

    # --- Raw data sheet ---
    if template.show_raw_data:
        ws_raw = wb.create_sheet(title=L["raw_sheet"])
        _header_row(ws_raw, [L["tag"], L["time"], L["value"], L["quality"]], row=1)
        row_idx = 2
        for td in per_tag_data:
            tag = td["tag"]
            for ts, val, qual in td.get("raw_readings", []):
                ws_raw.cell(row=row_idx, column=1, value=tag.name)
                ws_raw.cell(row=row_idx, column=2, value=ts.isoformat())
                c = ws_raw.cell(
                    row=row_idx, column=3, value=round(val, 3) if val is not None else None
                )
                c.number_format = FLOAT_FMT
                ws_raw.cell(row=row_idx, column=4, value=qual)
                row_idx += 1

    # --- Grafana panels sheet ---
    gf = [g for g in (grafana_charts or [])]
    if gf:
        ws_gf = wb.create_sheet(title="Grafana")
        row = 1
        for gc in gf:
            ws_gf.cell(row=row, column=1, value=gc["title"]).font = HEADER_FONT
            row += 1
            if gc.get("png"):
                _embed_image(ws_gf, gc["png"], f"A{row}")
                row += 15  # reserve rows for image
            else:
                ws_gf.cell(row=row, column=1, value=gc.get("error") or "render edilemedi")
                row += 2

    # --- Tesis Değişkenleri sayfası (yalnızca değişken varsa eklenir) ---
    if variables:
        vws = wb.create_sheet("Tesis Değişkenleri")
        _header_row(vws, ["Kod", "Ad", "Birim", "Tür", "Değer / Seri", "Uyarı"], row=1)
        r = 2
        for v in variables:
            if v["kind"] == "scalar":
                val_str = "" if v["value"] is None else f"{v['value']}"
            else:
                pts = v.get("points") or []
                val_str = f"{len(pts)} nokta"
            vws.cell(row=r, column=1, value=v["code"])
            vws.cell(row=r, column=2, value=v["name"])
            vws.cell(row=r, column=3, value=v["unit"])
            vws.cell(row=r, column=4, value=v["kind"])
            vws.cell(row=r, column=5, value=val_str)
            vws.cell(row=r, column=6, value=v.get("warning") or "")
            r += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
