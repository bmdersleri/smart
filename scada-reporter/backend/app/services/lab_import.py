import csv
import io


def parse_table(content: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    """Parse a CSV or XLSX upload into (headers, rows-of-strings).

    Empty trailing cells are normalized to "". The first row is the header.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = list(csv.reader(io.StringIO(text)))
        if not reader:
            return [], []
        headers = [h.strip() for h in reader[0]]
        rows = [[(c or "").strip() for c in r] for r in reader[1:] if any(c.strip() for c in r)]
        return headers, rows
    if name.endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        all_rows = [
            [("" if c is None else str(c)).strip() for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        if not all_rows:
            return [], []
        headers = all_rows[0]
        rows = [r for r in all_rows[1:] if any(c for c in r)]
        return headers, rows
    raise ValueError("Desteklenmeyen dosya turu (.csv veya .xlsx)")
