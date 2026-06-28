"""Compliance report-pack generation service.

Assembles the period-level official report-pack data from the compliance
models + engine events + notes, and renders it to JSON / Excel / PDF. The
three renderers all consume the same ``build_report_pack_data`` dict so the
logical sections stay identical across formats (per the design's
"Report Pack Content" section).
"""

from __future__ import annotations

import json
import os
from datetime import datetime
from io import BytesIO

from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from weasyprint import HTML

from app.models.compliance import (
    ComplianceDischargePoint,
    ComplianceEvent,
    ComplianceEventNote,
    ComplianceLimit,
    ComplianceParameter,
    CompliancePermit,
)

# Logical section keys present in every report pack (JSON/Excel/PDF parity).
SECTION_KEYS = (
    "cover",
    "period_summary",
    "parameter_limit_table",
    "measurement_results",
    "event_summary",
    "missing_sample_list",
    "bad_quality_list",
    "operator_explanations",
    "approval_block",
    "audit_metadata",
)

_template_dir = os.path.join(os.path.dirname(__file__), "..", "templates")
_env = Environment(loader=FileSystemLoader(_template_dir), autoescape=True)

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _iso(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.isoformat()


def _evidence(event: ComplianceEvent) -> dict:
    try:
        return json.loads(event.evidence_json) if event.evidence_json else {}
    except ValueError, TypeError:
        return {}


async def build_report_pack_data(
    db: AsyncSession,
    permit_id: int,
    period_start: datetime,
    period_end: datetime,
) -> dict:
    """Assemble the report-pack data dict for a permit + period.

    Pulls permit metadata, discharge points, parameters + limits, the durable
    compliance events for the period, and operator notes, and groups them into
    the logical report sections. Read-only: does not mutate any rows.
    """
    permit = await db.get(CompliancePermit, permit_id)

    points = list(
        (
            await db.execute(
                select(ComplianceDischargePoint)
                .where(ComplianceDischargePoint.permit_id == permit_id)
                .order_by(ComplianceDischargePoint.id)
            )
        )
        .scalars()
        .all()
    )
    point_by_id = {p.id: p for p in points}

    parameters = list(
        (
            await db.execute(
                select(ComplianceParameter)
                .where(ComplianceParameter.permit_id == permit_id)
                .order_by(ComplianceParameter.id)
            )
        )
        .scalars()
        .all()
    )
    param_by_id = {p.id: p for p in parameters}
    param_ids = [p.id for p in parameters]

    limits: list[ComplianceLimit] = []
    if param_ids:
        limits = list(
            (
                await db.execute(
                    select(ComplianceLimit)
                    .where(ComplianceLimit.parameter_id.in_(param_ids))
                    .order_by(ComplianceLimit.id)
                )
            )
            .scalars()
            .all()
        )

    events = list(
        (
            await db.execute(
                select(ComplianceEvent)
                .where(
                    ComplianceEvent.permit_id == permit_id,
                    ComplianceEvent.period_start == period_start.replace(tzinfo=None),
                    ComplianceEvent.period_end == period_end.replace(tzinfo=None),
                )
                .order_by(ComplianceEvent.id)
            )
        )
        .scalars()
        .all()
    )
    event_ids = [e.id for e in events]

    notes: list[ComplianceEventNote] = []
    if event_ids:
        notes = list(
            (
                await db.execute(
                    select(ComplianceEventNote)
                    .where(ComplianceEventNote.event_id.in_(event_ids))
                    .order_by(ComplianceEventNote.id)
                )
            )
            .scalars()
            .all()
        )

    # --- cover ---------------------------------------------------------------
    cover = {
        "permit_id": permit_id,
        "permit_name": permit.name if permit else None,
        "facility_name": permit.facility_name if permit else "",
        "authority": permit.authority if permit else "",
        "permit_number": permit.permit_number if permit else "",
        "report_frequency": permit.report_frequency if permit else "",
        "valid_from": _iso(permit.valid_from) if permit else None,
        "valid_to": _iso(permit.valid_to) if permit else None,
    }

    # --- period_summary ------------------------------------------------------
    by_type: dict[str, int] = {}
    by_status: dict[str, int] = {}
    for ev in events:
        by_type[ev.event_type] = by_type.get(ev.event_type, 0) + 1
        by_status[ev.status] = by_status.get(ev.status, 0) + 1
    period_summary = {
        "period_start": _iso(period_start),
        "period_end": _iso(period_end),
        "total_events": len(events),
        "open_events": by_status.get("open", 0),
        "by_event_type": by_type,
        "by_status": by_status,
        "parameter_count": len(parameters),
        "discharge_point_count": len(points),
    }

    # --- parameter_limit_table ----------------------------------------------
    parameter_limit_table = []
    for param in parameters:
        point = point_by_id.get(param.discharge_point_id)
        for limit in [limit for limit in limits if limit.parameter_id == param.id]:
            parameter_limit_table.append(
                {
                    "parameter_id": param.id,
                    "parameter_name": param.parameter_name,
                    "unit": param.unit,
                    "source_type": param.source_type,
                    "discharge_point": point.code if point else None,
                    "limit_id": limit.id,
                    "limit_type": limit.limit_type,
                    "min_value": limit.min_value,
                    "max_value": limit.max_value,
                    "aggregation": limit.aggregation,
                    "severity": limit.severity,
                    "requires_explanation": limit.requires_explanation,
                }
            )

    # --- measurement_results -------------------------------------------------
    measurement_results = []
    for ev in events:
        ev_param = param_by_id.get(ev.parameter_id)
        measurement_results.append(
            {
                "event_id": ev.id,
                "parameter_id": ev.parameter_id,
                "parameter_name": ev_param.parameter_name if ev_param else None,
                "event_type": ev.event_type,
                "observed_value": ev.observed_value,
                "limit_value": ev.limit_value,
                "status": ev.status,
            }
        )

    # --- event_summary -------------------------------------------------------
    event_summary = []
    for ev in events:
        ev_param = param_by_id.get(ev.parameter_id)
        event_summary.append(
            {
                "event_id": ev.id,
                "event_key": ev.event_key,
                "parameter_id": ev.parameter_id,
                "parameter_name": ev_param.parameter_name if ev_param else None,
                "limit_id": ev.limit_id,
                "event_type": ev.event_type,
                "severity": ev.severity,
                "status": ev.status,
                "observed_value": ev.observed_value,
                "limit_value": ev.limit_value,
                "evidence": _evidence(ev),
            }
        )

    missing_sample_list = [e for e in event_summary if e["event_type"] == "missing_sample"]
    bad_quality_list = [e for e in event_summary if e["event_type"] == "bad_quality"]

    # --- operator_explanations ----------------------------------------------
    operator_explanations = [
        {
            "note_id": note.id,
            "event_id": note.event_id,
            "user_id": note.user_id,
            "note": note.note,
            "created_at": _iso(note.created_at),
        }
        for note in notes
    ]

    # --- approval_block ------------------------------------------------------
    approval_block = {
        "status": "draft",
        "prepared_by": None,
        "approved_by": None,
        "approved_at": None,
    }

    # --- audit_metadata ------------------------------------------------------
    audit_metadata = {
        "generated_at": _iso(datetime.utcnow()),
        "permit_id": permit_id,
        "event_count": len(events),
        "note_count": len(notes),
    }

    return {
        "cover": cover,
        "period_summary": period_summary,
        "parameter_limit_table": parameter_limit_table,
        "measurement_results": measurement_results,
        "event_summary": event_summary,
        "missing_sample_list": missing_sample_list,
        "bad_quality_list": bad_quality_list,
        "operator_explanations": operator_explanations,
        "approval_block": approval_block,
        "audit_metadata": audit_metadata,
    }


def render_json(data: dict) -> bytes:
    """Compact JSON of all report-pack sections."""
    return json.dumps(data, separators=(",", ":"), sort_keys=True, default=str).encode("utf-8")


def _write_header(ws, cols: list[str], row: int = 1) -> None:
    for col_idx, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def render_excel(data: dict) -> bytes:
    """Sectioned openpyxl workbook — one sheet per major section."""
    wb = Workbook()

    # Cover sheet
    ws = wb.active
    ws.title = "Cover"
    ws.cell(row=1, column=1, value="Compliance Report Pack").font = _HEADER_FONT
    row = 3
    for key, value in data["cover"].items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    # Period summary sheet
    ws_ps = wb.create_sheet(title="Period Summary")
    ps = data["period_summary"]
    row = 1
    for key, value in ps.items():
        ws_ps.cell(row=row, column=1, value=key)
        ws_ps.cell(row=row, column=2, value=json.dumps(value) if isinstance(value, dict) else value)
        row += 1

    # Parameter / limit table
    ws_pl = wb.create_sheet(title="Limits")
    pl_cols = [
        "parameter_name",
        "unit",
        "source_type",
        "discharge_point",
        "limit_type",
        "min_value",
        "max_value",
        "aggregation",
        "severity",
        "requires_explanation",
    ]
    _write_header(ws_pl, pl_cols)
    for r, item in enumerate(data["parameter_limit_table"], start=2):
        for c, key in enumerate(pl_cols, start=1):
            ws_pl.cell(row=r, column=c, value=item.get(key))

    # Measurement results
    ws_mr = wb.create_sheet(title="Measurements")
    mr_cols = [
        "parameter_name",
        "event_type",
        "observed_value",
        "limit_value",
        "status",
    ]
    _write_header(ws_mr, mr_cols)
    for r, item in enumerate(data["measurement_results"], start=2):
        for c, key in enumerate(mr_cols, start=1):
            ws_mr.cell(row=r, column=c, value=item.get(key))

    # Event summary
    ws_es = wb.create_sheet(title="Events")
    es_cols = [
        "parameter_name",
        "event_type",
        "severity",
        "status",
        "observed_value",
        "limit_value",
    ]
    _write_header(ws_es, es_cols)
    for r, item in enumerate(data["event_summary"], start=2):
        for c, key in enumerate(es_cols, start=1):
            ws_es.cell(row=r, column=c, value=item.get(key))

    # Missing samples + bad quality
    ws_ms = wb.create_sheet(title="Missing Samples")
    _write_header(ws_ms, ["parameter_name", "observed_value", "limit_value", "status"])
    for r, item in enumerate(data["missing_sample_list"], start=2):
        ws_ms.cell(row=r, column=1, value=item.get("parameter_name"))
        ws_ms.cell(row=r, column=2, value=item.get("observed_value"))
        ws_ms.cell(row=r, column=3, value=item.get("limit_value"))
        ws_ms.cell(row=r, column=4, value=item.get("status"))

    ws_bq = wb.create_sheet(title="Bad Quality")
    _write_header(ws_bq, ["parameter_name", "observed_value", "limit_value", "status"])
    for r, item in enumerate(data["bad_quality_list"], start=2):
        ws_bq.cell(row=r, column=1, value=item.get("parameter_name"))
        ws_bq.cell(row=r, column=2, value=item.get("observed_value"))
        ws_bq.cell(row=r, column=3, value=item.get("limit_value"))
        ws_bq.cell(row=r, column=4, value=item.get("status"))

    # Operator explanations
    ws_oe = wb.create_sheet(title="Explanations")
    _write_header(ws_oe, ["event_id", "user_id", "note", "created_at"])
    for r, item in enumerate(data["operator_explanations"], start=2):
        ws_oe.cell(row=r, column=1, value=item.get("event_id"))
        ws_oe.cell(row=r, column=2, value=item.get("user_id"))
        ws_oe.cell(row=r, column=3, value=item.get("note"))
        ws_oe.cell(row=r, column=4, value=item.get("created_at"))

    # Approval + audit
    ws_ap = wb.create_sheet(title="Approval")
    row = 1
    for key, value in {**data["approval_block"], **data["audit_metadata"]}.items():
        ws_ap.cell(row=row, column=1, value=key)
        ws_ap.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def render_pdf(data: dict, lang: str = "en") -> bytes:
    """Render the report pack to PDF via the jinja2 + weasyprint template."""
    html_str = _env.get_template("compliance_report.html.j2").render(data=data, lang=lang)
    return HTML(string=html_str).write_pdf()


__all__ = [
    "SECTION_KEYS",
    "build_report_pack_data",
    "render_excel",
    "render_json",
    "render_pdf",
]
