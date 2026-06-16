"""Excel şablonu otomatik analiz: sayfa, başlık satırı, tarih sütunu, grid
başlangıcı tespit eder ve sensör kodlarından tag eşlemesi önerir."""

import re
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.tag import Tag

CODE_RE = re.compile(r"^\d{3}[A-Z]{2}\d{3}$")
_SCAN_ROWS = 6


_TR_FOLD = str.maketrans("İIŞÜÇÖĞ", "IISUCOG")


def _norm(s: object) -> str:
    # Tam Türkçe katlama: agg tahmin anahtar kelimeleri aksanlı harfte kaçmasın
    return str(s or "").strip().upper().translate(_TR_FOLD)


def _guess_agg(label: str) -> str:
    u = _norm(label)
    if "M3/GUN" in u or "DEBI" in u:
        return "sum"
    if "TUKETIM" in u or "SAYAC" in u:
        return "delta"
    if "%" in label or "ORAN" in u:
        return "avg"
    if "SEVIYE" in u:
        return "last"
    return "avg"


def _find_code_row(ws) -> int:
    best_row, best_hits = 1, -1
    for r in range(1, min(ws.max_row, _SCAN_ROWS) + 1):
        hits = sum(1 for c in ws[r] if isinstance(c.value, str) and CODE_RE.match(c.value.strip()))
        if hits > best_hits:
            best_row, best_hits = r, hits
    return best_row


def _find_date_col(ws, scan_rows: int) -> str:
    for r in range(1, scan_rows + 1):
        for c in ws[r]:
            if _norm(c.value) == "TARIH":
                return get_column_letter(c.column)
    return "A"


def _label_for(ws, col_idx: int, code_row: int) -> str:
    """Kod satırının üstündeki ilk dolu hücreyi etiket olarak al."""
    for r in range(code_row - 1, 0, -1):
        v = ws.cell(row=r, column=col_idx).value
        if v not in (None, ""):
            return str(v)
    return ""


async def inspect_template(db: AsyncSession, file_bytes: bytes) -> dict:
    wb = load_workbook(BytesIO(file_bytes), data_only=False)
    ws = wb.worksheets[0]
    code_row = _find_code_row(ws)
    date_col = _find_date_col(ws, min(ws.max_row, _SCAN_ROWS))

    codes: dict[int, str] = {}
    for c in ws[code_row]:
        if isinstance(c.value, str) and CODE_RE.match(c.value.strip()):
            codes[c.column] = c.value.strip()

    name_to_id: dict[str, int] = {}
    if codes:
        result = await db.execute(select(Tag.id, Tag.name).where(Tag.name.in_(codes.values())))
        name_to_id = {name: tid for tid, name in result.all()}

    columns = []
    for col_idx, code in sorted(codes.items()):
        label = _label_for(ws, col_idx, code_row)
        columns.append(
            {
                "col_letter": get_column_letter(col_idx),
                "source_code": code,
                "tag_id": name_to_id.get(code),
                "agg": _guess_agg(label),
                "label": label,
                "enabled": code in name_to_id,
            }
        )

    return {
        "sheet_name": ws.title,
        "header_row": code_row,
        "date_col": date_col,
        "data_start_row": code_row + 1,
        "date_mode": "write",
        "columns": columns,
    }
