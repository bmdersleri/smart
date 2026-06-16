"""Şablon kopyasına günlük değerleri yazıp xlsx bytes döndürür.

Temiz şablon (file_blob) yüklenir, seçilen ay için her eşlenmiş+aktif sütuna
günlük toplama yazılır. Verisi olmayan gün boş bırakılır (0 uydurma yok).
Hücre stili/format korunur — yalnız değer yazılır.
"""

import calendar
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import settings
from app.models.excel_template import ExcelTemplate
from app.services.template_fill.daily_rollup import daily_values


async def fill_template(db: AsyncSession, template_id: int, year: int, month: int) -> bytes:
    result = await db.execute(
        select(ExcelTemplate)
        .where(ExcelTemplate.id == template_id)
        .options(selectinload(ExcelTemplate.columns))
    )
    tpl = result.scalar_one_or_none()
    if tpl is None:
        raise ValueError(f"Şablon bulunamadı: {template_id}")

    wb = load_workbook(BytesIO(tpl.file_blob), data_only=False)
    ws = wb[tpl.sheet_name] if tpl.sheet_name in wb.sheetnames else wb.worksheets[0]
    offset = settings.REPORT_TZ_OFFSET_HOURS
    ndays = calendar.monthrange(year, month)[1]

    # write modunda gün -> satır eşlemesi; match modunda mevcut tarih hücreleri
    day_to_row: dict[int, int] = {}
    if tpl.date_mode == "match":
        for r in range(tpl.data_start_row, ws.max_row + 1):
            v = ws[f"{tpl.date_col}{r}"].value
            if isinstance(v, datetime):
                day_to_row[v.day] = r
    else:
        for day in range(1, ndays + 1):
            day_to_row[day] = tpl.data_start_row + (day - 1)

    # write modunda tarihleri yaz
    if tpl.date_mode == "write":
        for day in range(1, ndays + 1):
            ws[f"{tpl.date_col}{day_to_row[day]}"] = datetime(year, month, day)

    for col in tpl.columns:
        if not col.enabled or col.tag_id is None:
            continue
        vals = await daily_values(db, col.tag_id, year, month, col.agg, tz_offset_hours=offset)
        for day, value in vals.items():
            row = day_to_row.get(day)
            if row is not None:
                ws[f"{col.col_letter}{row}"] = value

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
