"""WinCC Excel export'larından tag kataloğu üretir.

- full_export.xlsx : tüm tag'ler (Name, Data type, Connection, Address) + Connections (IP)
- archive_export.xlsx : uzun-süre kaydı tutulan tag'ler (Tag name, Acquisition cycle) — adres yok
- gunluk_rapor.xlsx : günlük takip edilen cihaz kodları (best-effort)

archive Tag name -> full Name join ile mutlak adres + PLC IP + tip çözülür.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO

import openpyxl

_IP_RE = re.compile(r"S7ONLINE!::(\d+\.\d+\.\d+\.\d+):")
_INTERVAL_RE = re.compile(r"(\d+)\s*(second|minute|hour|saniye|dakika|saat)", re.IGNORECASE)
_TOKEN_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]{4,}$")


def parse_interval(value: object, default: int = 5) -> int:
    """'5 second' / '1 minute' / '2 hour' -> saniye."""
    if value is None:
        return default
    m = _INTERVAL_RE.search(str(value))
    if not m:
        return default
    n = int(m.group(1))
    unit = m.group(2).lower()
    if unit in ("minute", "dakika"):
        return n * 60
    if unit in ("hour", "saat"):
        return n * 3600
    return n


def _open(source: str | bytes) -> openpyxl.Workbook:
    if isinstance(source, bytes):
        return openpyxl.load_workbook(BytesIO(source), read_only=True, data_only=True)
    return openpyxl.load_workbook(source, read_only=True, data_only=True)


def load_connections(wb: openpyxl.Workbook) -> dict[str, str]:
    """Connections sayfası: bağlantı adı -> IP."""
    if "Connections" not in wb.sheetnames:
        return {}
    ws = wb["Connections"]
    out: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not row or row[0] is None:
            continue
        params = row[3] if len(row) > 3 else None
        m = _IP_RE.search(str(params)) if params else None
        if m:
            out[str(row[0])] = m.group(1)
    return out


def load_full_tags(wb: openpyxl.Workbook) -> dict[str, tuple[str, str, str]]:
    """Tags sayfası: Name -> (connection, address, data_type)."""
    ws = wb["Tags"]
    out: dict[str, tuple[str, str, str]] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not row or row[0] is None:
            continue
        name = str(row[0])
        dtype = str(row[2]) if len(row) > 2 and row[2] else ""
        conn = str(row[5]) if len(row) > 5 and row[5] else ""
        addr = str(row[7]) if len(row) > 7 and row[7] else ""
        out[name] = (conn, addr, dtype)
    return out


def load_archive_tags(wb: openpyxl.Workbook) -> list[tuple[str, int, bool]]:
    """archive Tags sayfası: (tag_name, interval_seconds, long_term)."""
    ws = wb["Tags"]
    out: list[tuple[str, int, bool]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not row or row[0] is None:
            continue
        name = str(row[2]) if len(row) > 2 and row[2] else None
        if not name:
            continue
        interval = parse_interval(row[9] if len(row) > 9 else None)
        long_term = bool(row[25]) if len(row) > 25 and row[25] is not None else True
        out.append((name, interval, long_term))
    return out


@dataclass
class CatalogResult:
    tags: list[dict] = field(default_factory=list)
    resolved: int = 0
    skipped: int = 0  # archive'da var ama full'de adres yok (SiproTec/UMG)
    skipped_names: list[str] = field(default_factory=list)
    daily_matched: int = 0
    daily_unmatched: list[str] = field(default_factory=list)


def _tag_dict(
    name: str, conn: str, addr: str, dtype: str, ip: str, interval: int, long_term: bool
) -> dict:
    return {
        "node_id": name,
        "name": name,
        "description": f"{conn} — {ip}" if ip else conn,
        "unit": "",
        "channel": "",
        "device": conn,
        "plc_name": conn,
        "plc_ip": ip or None,
        "plc_rack": 0,
        "plc_slot": 1,
        "s7_address": addr or None,
        "data_type": dtype,
        "sample_interval": interval,
        "long_term": long_term,
        "daily_tracking": False,
    }


def build_archive_catalog(
    full_source: str | bytes,
    archive_source: str | bytes,
    gunluk_source: str | bytes | None = None,
) -> CatalogResult:
    """archive tag'lerini full_export ile join ederek katalog üretir."""
    full_wb = _open(full_source)
    conn_ip = load_connections(full_wb)
    full_tags = load_full_tags(full_wb)
    full_wb.close()

    arch_wb = _open(archive_source)
    archive = load_archive_tags(arch_wb)
    arch_wb.close()

    res = CatalogResult()
    for name, interval, long_term in archive:
        if name not in full_tags:
            res.skipped += 1
            if len(res.skipped_names) < 50:
                res.skipped_names.append(name)
            continue
        conn, addr, dtype = full_tags[name]
        ip = conn_ip.get(conn, "")
        res.tags.append(_tag_dict(name, conn, addr, dtype, ip, interval, long_term))
        res.resolved += 1

    if gunluk_source is not None:
        _apply_daily_tracking(res, full_tags, gunluk_source)

    return res


def _apply_daily_tracking(
    res: CatalogResult,
    full_tags: dict[str, tuple[str, str, str]],
    gunluk_source: str | bytes,
) -> None:
    """gunluk_rapor token'ları -> 'B110<token>' ile başlayan full tag'leri eşle.

    Best-effort: token cihaz kodudur, birden çok tag'e karşılık gelebilir.
    Katalogda bulunan eşleşmeler daily_tracking=True işaretlenir; eşleşmeyen
    token'lar rapor edilir (otomatik import edilmez — son ek belirsiz).
    """
    tokens = _gunluk_tokens(gunluk_source)
    by_name = {t["name"]: t for t in res.tags}
    # full tag adlarını prefix aramaya hazırla
    full_names = list(full_tags.keys())
    for token in sorted(tokens):
        prefix = f"B110{token}"
        matches = [n for n in full_names if n.startswith(prefix)]
        in_catalog = [n for n in matches if n in by_name]
        if in_catalog:
            for n in in_catalog:
                if not by_name[n]["daily_tracking"]:
                    by_name[n]["daily_tracking"] = True
                    res.daily_matched += 1
        if not matches and len(res.daily_unmatched) < 100:
            res.daily_unmatched.append(token)


def _gunluk_tokens(source: str | bytes) -> set[str]:
    wb = _open(source)
    tokens: set[str] = set()
    for ws in wb.worksheets:
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 30:  # tag kodları üst başlık satırlarında
                break
            for c in row:
                if isinstance(c, str):
                    s = c.strip()
                    if _TOKEN_RE.match(s) and any(ch.isdigit() for ch in s):
                        tokens.add(s)
    wb.close()
    return tokens


def build_full_catalog(full_source: str | bytes) -> CatalogResult:
    """full_export'taki TÜM tag'leri (adres+IP çözülmüş) katalog yapar (UI import)."""
    full_wb = _open(full_source)
    conn_ip = load_connections(full_wb)
    full_tags = load_full_tags(full_wb)
    full_wb.close()

    res = CatalogResult()
    for name, (conn, addr, dtype) in full_tags.items():
        if conn == "Internal tags" or not addr:
            res.skipped += 1
            continue
        ip = conn_ip.get(conn, "")
        res.tags.append(_tag_dict(name, conn, addr, dtype, ip, 5, False))
        res.resolved += 1
    return res
