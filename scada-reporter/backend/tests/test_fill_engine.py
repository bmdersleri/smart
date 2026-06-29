from datetime import datetime
from io import BytesIO

import pytest
import pytest_asyncio
from openpyxl import Workbook, load_workbook

from app.models.excel_template import ExcelTemplate, ExcelTemplateColumn
from app.models.tag import Tag, TagReading
from app.services.template_fill.fill_engine import fill_template


@pytest_asyncio.fixture(autouse=True)
async def _clean(db_session):
    yield
    from sqlalchemy import delete

    await db_session.execute(delete(ExcelTemplateColumn))
    await db_session.execute(delete(ExcelTemplate))
    await db_session.execute(delete(TagReading))
    await db_session.execute(delete(Tag))
    await db_session.commit()


def _template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "OCAK 2026"
    ws["D2"] = "SENSÖR KODLARI"
    ws["E2"] = "410BF103"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest_asyncio.fixture
async def saved_template(db_session):
    tag = Tag(node_id="a", name="410BF103", unit="m3")
    db_session.add(tag)
    await db_session.flush()
    db_session.add_all(
        [
            TagReading(tag_id=tag.id, value=10.0, timestamp=datetime(2026, 1, 1, 6, 0)),
            TagReading(tag_id=tag.id, value=30.0, timestamp=datetime(2026, 1, 1, 18, 0)),
            TagReading(tag_id=tag.id, value=50.0, timestamp=datetime(2026, 1, 3, 9, 0)),
        ]
    )
    tpl = ExcelTemplate(
        name="T",
        file_blob=_template_bytes(),
        sheet_name="OCAK 2026",
        header_row=2,
        date_col="D",
        data_start_row=3,
        date_mode="write",
    )
    tpl.columns = [
        ExcelTemplateColumn(col_letter="E", tag_id=tag.id, agg="sum", source_code="410BF103"),
    ]
    db_session.add(tpl)
    await db_session.commit()
    await db_session.refresh(tpl)
    return tpl


@pytest.mark.asyncio
async def test_fill_writes_daily_sums_and_dates(db_session, saved_template):
    out = await fill_template(db_session, saved_template.id, 2026, 1)
    ws = load_workbook(BytesIO(out)).active
    assert ws["E3"].value == 40.0  # 10 + 30 (day 1)
    assert ws["E5"].value == 50.0  # day 3
    assert ws["E4"].value is None  # day 2 no data -> blank, not 0
    assert ws["D3"].value.day == 1
    assert ws["D5"].value.day == 3


@pytest.mark.asyncio
async def test_disabled_and_null_columns_skipped(db_session, saved_template):
    saved_template.columns[0].enabled = False
    await db_session.commit()
    out = await fill_template(db_session, saved_template.id, 2026, 1)
    ws = load_workbook(BytesIO(out)).active
    assert ws["E3"].value is None


@pytest.mark.asyncio
async def test_write_mode_dates_get_turkish_format(db_session, saved_template):
    out = await fill_template(db_session, saved_template.id, 2026, 1)
    ws = load_workbook(BytesIO(out)).active
    # Boş tarih sütununa yazılan tarih ISO değil TR biçiminde olmalı
    assert ws["D3"].number_format == "dd.mm.yyyy"


def _match_template_bytes() -> bytes:
    """match modu: tarih hücreleri önceden dolu, satırlar bitişik değil."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OCAK 2026"
    ws["D2"] = "SENSÖR KODLARI"
    ws["E2"] = "410BF103"
    ws["D3"] = datetime(2026, 1, 2)  # day 2 -> row 3
    ws["D4"] = datetime(2026, 1, 5)  # day 5 -> row 4 (day 3 has NO date cell)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.mark.asyncio
async def test_match_mode_maps_existing_date_cells(db_session):
    tag = Tag(node_id="a", name="410BF103", unit="m3")
    db_session.add(tag)
    await db_session.flush()
    db_session.add_all(
        [
            TagReading(tag_id=tag.id, value=12.0, timestamp=datetime(2026, 1, 2, 6, 0)),
            # day 3 has no date cell in the match template -> must be skipped
            TagReading(tag_id=tag.id, value=99.0, timestamp=datetime(2026, 1, 3, 6, 0)),
            TagReading(tag_id=tag.id, value=7.0, timestamp=datetime(2026, 1, 5, 6, 0)),
        ]
    )
    tpl = ExcelTemplate(
        name="M",
        file_blob=_match_template_bytes(),
        sheet_name="OCAK 2026",
        header_row=2,
        date_col="D",
        data_start_row=3,
        date_mode="match",
    )
    tpl.columns = [
        ExcelTemplateColumn(col_letter="E", tag_id=tag.id, agg="sum", source_code="410BF103"),
    ]
    db_session.add(tpl)
    await db_session.commit()
    await db_session.refresh(tpl)

    out = await fill_template(db_session, tpl.id, 2026, 1)
    ws = load_workbook(BytesIO(out)).active
    assert ws["E3"].value == 12.0  # day 2 mapped to its existing date row
    assert ws["E4"].value == 7.0  # day 5
    # day 3 (value 99) has no date cell in the template -> silently skipped,
    # and it must NOT leak into any other row
    assert ws["E5"].value is None


@pytest.mark.asyncio
async def test_null_tag_column_skipped(db_session, saved_template):
    # tag_id None (unmapped/manual column) -> skipped even though enabled
    saved_template.columns[0].tag_id = None
    await db_session.commit()
    out = await fill_template(db_session, saved_template.id, 2026, 1)
    ws = load_workbook(BytesIO(out)).active
    assert ws["E3"].value is None


@pytest.mark.asyncio
async def test_missing_template_raises(db_session):
    with pytest.raises(ValueError):
        await fill_template(db_session, 9999, 2026, 1)


@pytest.mark.asyncio
async def test_fill_variable_series_column(db_session):
    from io import BytesIO

    from openpyxl import load_workbook

    from app.models.excel_template import ExcelTemplate, ExcelTemplateColumn
    from app.models.tag import Tag, TagReading
    from app.services.facility_variables.service import create_variable
    from app.services.template_fill.fill_engine import fill_template

    tag = Tag(node_id="n", name="T", unit="m3")
    db_session.add(tag)
    await db_session.commit()
    await db_session.refresh(tag)
    # UTC 06:00 ve 18:00 → UTC+3 yerel saatte 09:00 ve 21:00, aynı gün bucketı
    db_session.add(TagReading(tag_id=tag.id, timestamp=datetime(2026, 6, 1, 6), value=0.0))
    db_session.add(TagReading(tag_id=tag.id, timestamp=datetime(2026, 6, 1, 18), value=40.0))
    await db_session.commit()

    var = await create_variable(
        db_session,
        code="v_fill",
        name="v",
        description="",
        kind="series",
        unit="m3/gun",
        expression={
            "op": "series",
            "source": {"type": "tag", "tag_id": tag.id},
            "agg": "delta",
            "grain": "day",
            "window": "month",
        },
        null_policy="skip",
        quality_policy="good_only",
        default_time_grain="day",
        value_type="number",
        created_by=1,
    )

    wb = Workbook()  # Workbook üst seviyede import edilmiş
    ws = wb.active
    ws.title = "S"
    buf = BytesIO()
    wb.save(buf)

    tpl = ExcelTemplate(
        name="vf",
        description="",
        file_blob=buf.getvalue(),
        sheet_name="S",
        header_row=1,
        date_col="A",
        data_start_row=2,
        date_mode="write",
    )
    tpl.columns = [
        ExcelTemplateColumn(
            col_letter="K",
            source_type="variable",
            variable_id=var.id,
            write_mode="series",
            target_mode="column",
            enabled=True,
        )
    ]
    db_session.add(tpl)
    await db_session.commit()
    await db_session.refresh(tpl)

    out = await fill_template(db_session, tpl.id, 2026, 6)
    rwb = load_workbook(BytesIO(out))
    rws = rwb["S"]
    # gün 1 → data_start_row (2): K2 == 40.0
    assert rws["K2"].value == 40.0


@pytest.mark.asyncio
async def test_fill_variable_reduce_cell(db_session):
    from io import BytesIO

    from openpyxl import load_workbook

    from app.models.excel_template import ExcelTemplate, ExcelTemplateColumn
    from app.models.tag import Tag, TagReading
    from app.services.facility_variables.service import create_variable
    from app.services.template_fill.fill_engine import fill_template

    tag = Tag(node_id="n", name="T", unit="m3")
    db_session.add(tag)
    await db_session.commit()
    await db_session.refresh(tag)
    # UTC 06:00 ve 18:00 → UTC+3 yerel saatte 09:00 ve 21:00, aynı gün bucketı
    for ts, v in (
        (datetime(2026, 6, 1, 6), 0.0),
        (datetime(2026, 6, 1, 18), 40.0),
        (datetime(2026, 6, 2, 6), 40.0),
        (datetime(2026, 6, 2, 18), 80.0),
    ):
        db_session.add(TagReading(tag_id=tag.id, timestamp=ts, value=v))
    await db_session.commit()

    var = await create_variable(
        db_session,
        code="v_red",
        name="v",
        description="",
        kind="series",
        unit="m3/gun",
        expression={
            "op": "series",
            "source": {"type": "tag", "tag_id": tag.id},
            "agg": "delta",
            "grain": "day",
            "window": "month",
        },
        null_policy="skip",
        quality_policy="good_only",
        default_time_grain="day",
        value_type="number",
        created_by=1,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    buf = BytesIO()
    wb.save(buf)

    tpl = ExcelTemplate(
        name="vr",
        description="",
        file_blob=buf.getvalue(),
        sheet_name="S",
        header_row=1,
        date_col="A",
        data_start_row=2,
        date_mode="write",
    )
    tpl.columns = [
        ExcelTemplateColumn(
            col_letter="K",
            source_type="variable",
            variable_id=var.id,
            write_mode="reduce",
            reduce_op="avg",
            target_mode="cell",
            target_cell="M5",
            enabled=True,
        )
    ]
    db_session.add(tpl)
    await db_session.commit()
    await db_session.refresh(tpl)

    out = await fill_template(db_session, tpl.id, 2026, 6)
    rwb = load_workbook(BytesIO(out))
    rws = rwb["S"]
    assert rws["M5"].value == 40.0  # avg(40, 40)


@pytest.mark.asyncio
async def test_variable_scalar_no_target_cell_logs_warning(db_session, caplog):
    """FIX 1 regresyon: target_cell=None olan hücre-hedefli değer sessizce düşürülmemeli,
    uyarı loglanmalı ve çalışma kitabına stray değer yazılmamalı."""
    import logging

    from app.models.excel_template import ExcelTemplate, ExcelTemplateColumn
    from app.services.facility_variables.service import create_variable
    from app.services.template_fill.fill_engine import fill_template

    # Sabit 42.0 üreten scalar değişken (const op → kind="scalar")
    var = await create_variable(
        db_session,
        code="v_notarget",
        name="notarget",
        description="",
        kind="scalar",
        unit="",
        expression={"op": "const", "value": 42.0},
        null_policy="skip",
        quality_policy="good_only",
        default_time_grain="day",
        value_type="number",
        created_by=1,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    buf = BytesIO()
    wb.save(buf)

    tpl = ExcelTemplate(
        name="vnt",
        description="",
        file_blob=buf.getvalue(),
        sheet_name="S",
        header_row=1,
        date_col="A",
        data_start_row=2,
        date_mode="write",
    )
    # target_cell=None ile target_mode="column": scalar değer üretilir ama nereye yazılacağı yok
    tpl.columns = [
        ExcelTemplateColumn(
            col_letter="Z",
            source_type="variable",
            variable_id=var.id,
            write_mode="reduce",
            reduce_op="avg",
            target_mode="column",
            target_cell=None,
            enabled=True,
        )
    ]
    db_session.add(tpl)
    await db_session.commit()
    await db_session.refresh(tpl)

    with caplog.at_level(logging.WARNING, logger="app.services.template_fill.fill_engine"):
        out = await fill_template(db_session, tpl.id, 2026, 1)

    # Uyarı loglanmış olmalı ve kolon harfini içermeli
    warning_msgs = [r.message for r in caplog.records if r.levelno == logging.WARNING]
    assert any("Z" in m for m in warning_msgs), (
        f"'Z' sütunu için uyarı logu bekleniyor ama bulunamadı. Loglar: {warning_msgs}"
    )

    # Çalışma kitabında Z sütununa stray değer yazılmamış olmalı
    rws = load_workbook(BytesIO(out)).active
    for row in range(1, rws.max_row + 1):
        cell_val = rws[f"Z{row}"].value
        assert cell_val is None, f"Z{row} beklenmedik değer içeriyor: {cell_val}"
