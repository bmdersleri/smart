from io import BytesIO

import pytest
import pytest_asyncio
from openpyxl import Workbook
from sqlalchemy import delete

from app.models.tag import Tag
from app.services.template_fill.template_inspector import inspect_template


@pytest_asyncio.fixture(autouse=True)
async def _clean_tags(db_session):
    yield
    await db_session.execute(delete(Tag))
    await db_session.commit()


def _make_template_bytes() -> bytes:
    """row1 başlık metni, row2 sensör kodları, row3+ tarih + boş grid."""
    wb = Workbook()
    ws = wb.active
    ws.title = "OCAK 2026"
    ws["D1"] = "TARİH"
    ws["E1"] = "TESİSE ALINAN DEBİ m3/gün"
    ws["F1"] = "ELEKTRİK TÜKETİMİ"
    ws["B1"] = "HAVA DURUMU"
    ws["D2"] = "SENSÖR KODLARI"
    ws["E2"] = "410BF103"
    ws["F2"] = "460BF105"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest_asyncio.fixture
async def seeded_tags(db_session):
    db_session.add_all(
        [
            Tag(node_id="a", name="410BF103", unit="m3"),
            Tag(node_id="b", name="460BF105", unit="kWh"),
        ]
    )
    await db_session.commit()


@pytest.mark.asyncio
async def test_detects_layout_and_mapping(db_session, seeded_tags):
    proposal = await inspect_template(db_session, _make_template_bytes())
    assert proposal["sheet_name"] == "OCAK 2026"
    assert proposal["header_row"] == 2
    assert proposal["date_col"] == "D"
    cols = {c["col_letter"]: c for c in proposal["columns"]}
    assert cols["E"]["source_code"] == "410BF103"
    assert cols["E"]["tag_id"] is not None
    assert cols["E"]["agg"] == "sum"
    assert cols["F"]["agg"] == "delta"
    assert cols["F"]["tag_id"] is not None


@pytest.mark.asyncio
async def test_unmatched_code_is_unmapped(db_session):
    proposal = await inspect_template(db_session, _make_template_bytes())
    cols = {c["col_letter"]: c for c in proposal["columns"]}
    assert cols["E"]["tag_id"] is None
    assert cols["E"]["source_code"] == "410BF103"
