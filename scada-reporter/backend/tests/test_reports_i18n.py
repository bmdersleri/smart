"""i18n tests for the legacy inline Excel export in app/api/reports.py (Task 14)."""

import io
from datetime import datetime

from openpyxl import load_workbook

from app.api.reports import _build_simple_excel


def _data() -> dict:
    return {
        "Pump1 (bar)": [
            {"period": "2026-06-15 00:00", "avg": 12.5, "min": 10.0, "max": 15.0, "count": 60},
        ]
    }


def _wb(lang: str):
    content = _build_simple_excel(
        _data(),
        start=datetime(2026, 6, 15, 0, 0),
        end=datetime(2026, 6, 15, 2, 0),
        interval="hourly",
        lang=lang,
    )
    return load_workbook(io.BytesIO(content))


def test_summary_sheet_title_english():
    wb = _wb("en")
    assert "Summary" in wb.sheetnames


def test_summary_sheet_title_turkish():
    wb = _wb("tr")
    assert "Özet" in wb.sheetnames


def test_header_row_differs_by_language():
    ws_en = _wb("en")["Pump1 (bar)"]
    ws_tr = _wb("tr")["Pump1 (bar)"]
    header_en = [c.value for c in ws_en[1]]
    header_tr = [c.value for c in ws_tr[1]]
    assert header_en[0] == "Period"
    assert header_tr[0] == "Dönem"
    assert header_en != header_tr


def test_report_title_localized():
    ws_en = _wb("en")["Summary"]
    ws_tr = _wb("tr")["Özet"]
    assert ws_en["A1"].value == "Report"
    assert ws_tr["A1"].value == "Rapor"
