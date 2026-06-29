from io import BytesIO
from types import SimpleNamespace

import pytest
import pytest_asyncio
from openpyxl import Workbook

from app.api.auth import get_current_user
from app.main import app
from app.models.tag import Tag


def _template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "OCAK 2026"
    ws["D2"] = "SENSÖR KODLARI"
    ws["E2"] = "410BF103"
    ws["E1"] = "DEBİ m3/gün"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest_asyncio.fixture(autouse=True)
def _auth_override():
    # Prod get_current_user, User ORM nesnesi döner (.id). dict değil obje ver.
    app.dependency_overrides[get_current_user] = lambda: SimpleNamespace(
        id=1, username="admin", role="admin", permission_overrides={}, is_active=True
    )
    yield
    app.dependency_overrides.pop(get_current_user, None)


@pytest_asyncio.fixture(autouse=True)
async def _clean(db_session):
    yield
    from sqlalchemy import delete

    from app.models.excel_template import ExcelTemplate, ExcelTemplateColumn

    await db_session.execute(delete(ExcelTemplateColumn))
    await db_session.execute(delete(ExcelTemplate))
    await db_session.execute(delete(Tag))
    await db_session.commit()


@pytest_asyncio.fixture
async def seeded_tag(db_session):
    tag = Tag(node_id="a", name="410BF103", unit="m3")
    db_session.add(tag)
    await db_session.commit()
    return tag


@pytest.mark.asyncio
async def test_inspect_returns_proposal(client, seeded_tag):
    resp = await client.post(
        "/api/excel-templates/inspect",
        files={
            "file": (
                "t.xlsx",
                _template_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["sheet_name"] == "OCAK 2026"
    assert data["columns"][0]["source_code"] == "410BF103"


@pytest.mark.asyncio
async def test_save_and_generate_roundtrip(client, seeded_tag):
    import base64

    payload = {
        "name": "Balta",
        "description": "",
        "file_b64": base64.b64encode(_template_bytes()).decode(),
        "sheet_name": "OCAK 2026",
        "header_row": 2,
        "date_col": "D",
        "data_start_row": 3,
        "date_mode": "write",
        "columns": [
            {
                "col_letter": "E",
                "tag_id": seeded_tag.id,
                "agg": "sum",
                "source_code": "410BF103",
                "enabled": True,
            }
        ],
    }
    save = await client.post("/api/excel-templates", json=payload)
    assert save.status_code == 201, save.text
    tpl_id = save.json()["id"]

    listed = await client.get("/api/excel-templates")
    assert any(t["id"] == tpl_id for t in listed.json())

    gen = await client.post(f"/api/excel-templates/{tpl_id}/generate?year=2026&month=1")
    assert gen.status_code == 200
    assert gen.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert gen.content[:2] == b"PK"


@pytest.mark.asyncio
async def test_duplicate_name_returns_409(client, seeded_tag):
    import base64

    payload = {
        "name": "Balta",
        "description": "",
        "file_b64": base64.b64encode(_template_bytes()).decode(),
        "sheet_name": "OCAK 2026",
        "header_row": 2,
        "date_col": "D",
        "data_start_row": 3,
        "date_mode": "write",
        "columns": [],
    }
    first = await client.post("/api/excel-templates", json=payload)
    assert first.status_code == 201, first.text
    dup = await client.post("/api/excel-templates", json=payload)
    assert dup.status_code == 409


@pytest.mark.asyncio
async def test_variable_binding_roundtrip(client, db_session):
    import base64

    from app.services.facility_variables.service import create_variable

    var = await create_variable(
        db_session,
        code="v_api",
        name="v",
        description="",
        kind="scalar",
        unit="m3/gun",
        expression={"op": "const", "value": 1.0},
        null_policy="skip",
        quality_policy="good_only",
        default_time_grain="day",
        value_type="number",
        created_by=1,
    )
    payload = {
        "name": "vbind-api",
        "description": "",
        "file_b64": base64.b64encode(_template_bytes()).decode(),
        "sheet_name": "OCAK 2026",
        "header_row": 2,
        "date_col": "D",
        "data_start_row": 3,
        "date_mode": "write",
        "columns": [
            {
                "col_letter": "K",
                "source_type": "variable",
                "variable_id": var.id,
                "write_mode": "reduce",
                "reduce_op": "sum",
                "target_mode": "cell",
                "target_cell": "K5",
                "variable_code_snapshot": "v_api",
                "enabled": True,
            }
        ],
    }
    resp = await client.post("/api/excel-templates", json=payload)
    assert resp.status_code == 201
    col = resp.json()["columns"][0]
    assert col["source_type"] == "variable"
    assert col["variable_id"] == var.id
    assert col["target_cell"] == "K5"


@pytest.mark.asyncio
async def test_variable_column_rejects_both_sources(client):
    import base64

    payload = {
        "name": "bad-both",
        "description": "",
        "file_b64": base64.b64encode(_template_bytes()).decode(),
        "sheet_name": "OCAK 2026",
        "header_row": 2,
        "date_col": "D",
        "data_start_row": 3,
        "date_mode": "write",
        "columns": [
            {
                "col_letter": "K",
                "source_type": "variable",
                "variable_id": 1,
                "tag_id": 9,
                "enabled": True,
            }
        ],
    }
    resp = await client.post("/api/excel-templates", json=payload)
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_generate_blocks_on_drift(client, seeded_tag):
    import base64

    # Map column E to source_code "999XX999", but the template blob's E2 cell
    # actually contains "410BF103" -> stored code != blob code -> drift.
    payload = {
        "name": "Drift",
        "description": "",
        "file_b64": base64.b64encode(_template_bytes()).decode(),
        "sheet_name": "OCAK 2026",
        "header_row": 2,
        "date_col": "D",
        "data_start_row": 3,
        "date_mode": "write",
        "columns": [
            {
                "col_letter": "E",
                "tag_id": seeded_tag.id,
                "agg": "sum",
                "source_code": "999XX999",
                "enabled": True,
            }
        ],
    }
    save = await client.post("/api/excel-templates", json=payload)
    assert save.status_code == 201, save.text
    tpl_id = save.json()["id"]
    gen = await client.post(f"/api/excel-templates/{tpl_id}/generate?year=2026&month=1")
    assert gen.status_code == 409
    assert "drift" in gen.json()["detail"].lower()
