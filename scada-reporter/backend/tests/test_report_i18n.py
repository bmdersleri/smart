"""i18n tests for the Excel report builder (Task 4)."""

import io
from dataclasses import dataclass, field
from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from app.services.excel_builder import build_advanced_excel
from app.services.pdf_builder import build_pdf
from app.services.stats_engine import TagStats


@dataclass
class _Tag:
    name: str = "Pump1"
    unit: str = "m3/h"


@dataclass
class _Template:
    show_summary_stats: bool = True
    show_anomaly_table: bool = True
    show_trend_charts: bool = False
    show_raw_data: bool = True


@dataclass
class _Archive:
    id: int = 1


def _stats() -> TagStats:
    return TagStats(
        tag_id=1,
        tag_name="Pump1",
        unit="m3/h",
        count=10,
        good_quality_count=10,
        availability_pct=100.0,
        avg=12.5,
        std_dev=1.2,
        variance=1.44,
        min=10.0,
        max=15.0,
        percentiles={10: 10.5, 50: 12.5, 90: 14.5},
        trend_slope=0.1,
        trend_r2=0.95,
        trend_direction="rising",
        rate_of_change_per_hour=0.1,
        gap_count=0,
        gap_total_seconds=0.0,
    )


@pytest.fixture
def sample_report_archive() -> _Archive:
    return _Archive()


@pytest.fixture
def sample_per_tag_data() -> list[dict]:
    return [
        {
            "tag": _Tag(),
            "stats": _stats(),
            "anomalies": [],
            "period_rows": [],
            "chart_png": b"",
            "raw_readings": [],
        }
    ]


def _build(archive, per_tag_data, lang: str) -> bytes:
    return build_advanced_excel(
        archive,
        per_tag_data,
        _Template(),
        b"",
        lang=lang,
    )


def _sheet_titles(content: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(content))
    return wb.sheetnames


def test_excel_uses_english_summary_sheet(sample_report_archive, sample_per_tag_data):
    content = _build(sample_report_archive, sample_per_tag_data, lang="en")
    assert "Summary" in _sheet_titles(content)


def test_excel_uses_turkish_summary_sheet(sample_report_archive, sample_per_tag_data):
    content = _build(sample_report_archive, sample_per_tag_data, lang="tr")
    assert "Özet" in _sheet_titles(content)


# ── PDF i18n (Task 5) ────────────────────────────────────────────────────────


@dataclass
class _PdfTemplate:
    name: str = "Daily Report"
    interval: str = "hourly"
    show_summary_stats: bool = True
    show_anomaly_table: bool = True
    show_trend_charts: bool = False


@dataclass
class _PdfArchive:
    id: int = 1
    start: datetime = field(default_factory=lambda: datetime(2026, 1, 1, 0, 0, tzinfo=UTC))
    end: datetime = field(default_factory=lambda: datetime(2026, 1, 2, 0, 0, tzinfo=UTC))


@pytest.fixture
def pdf_archive() -> _PdfArchive:
    return _PdfArchive()


def test_pdf_differs_by_language(pdf_archive, sample_per_tag_data):
    generated_at = datetime(2026, 1, 2, 12, 0, tzinfo=UTC)
    content_en = build_pdf(
        pdf_archive,
        sample_per_tag_data,
        _PdfTemplate(),
        "Test Facility",
        generated_at,
        lang="en",
    )
    content_tr = build_pdf(
        pdf_archive,
        sample_per_tag_data,
        _PdfTemplate(),
        "Test Facility",
        generated_at,
        lang="tr",
    )
    assert isinstance(content_en, (bytes, bytearray))
    assert content_en != content_tr  # localized labels produce different output
