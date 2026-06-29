"""report_archive stores resolved facility-variable refs; orchestrator renders variables."""

import gzip
import json
from dataclasses import dataclass
from datetime import UTC, datetime

import pytest

from app.models.report_archive import ReportArchive


@dataclass
class _VarTemplate:
    tag_ids: str = "[]"
    variable_ids: str = "[]"
    time_range_type: str = "custom"
    custom_start: datetime | None = None
    custom_end: datetime | None = None
    interval: str = "daily"
    output_format: str = "json"
    percentile_levels: str = "[50]"
    include_std_dev: bool = True
    include_percentiles: bool = True
    include_trend_line: bool = False
    anomaly_enabled: bool = False
    anomaly_zscore_threshold: float = 3.0
    show_summary_stats: bool = True
    show_trend_charts: bool = False
    show_anomaly_table: bool = False
    show_raw_data: bool = False
    grafana_panels: str = "[]"


async def _make_scalar_var(db, version=2):
    from app.models.facility_variable import FacilityVariable
    from app.models.tag import Tag, TagReading

    tag = Tag(node_id="ns=2;s=VG", name="VG", unit="m3")
    db.add(tag)
    await db.commit()
    await db.refresh(tag)
    db.add(TagReading(tag_id=tag.id, timestamp=datetime(2026, 6, 1, 1), value=10.0))
    db.add(TagReading(tag_id=tag.id, timestamp=datetime(2026, 6, 1, 20), value=70.0))
    await db.commit()
    var = FacilityVariable(
        code="var_orch",
        name="Orch Var",
        kind="scalar",
        unit="m3",
        version=version,
        default_time_grain="day",
        expression_json=json.dumps(
            {
                "op": "reduce",
                "reduce": "sum",
                "source": {
                    "op": "series",
                    "source": {"type": "tag", "tag_id": tag.id},
                    "agg": "delta",
                    "grain": "day",
                    "window": "day",
                },
            }
        ),
    )
    db.add(var)
    await db.commit()
    await db.refresh(var)
    return var


@pytest.mark.asyncio
async def test_orchestrator_stamps_variable_refs(db_session):
    from app.services.report_generator import generate_report_from_template

    var = await _make_scalar_var(db_session, version=7)
    arch = ReportArchive(
        status="pending",
        trigger="manual",
        tag_ids="[]",
        start=datetime(2026, 6, 1, tzinfo=UTC),
        end=datetime(2026, 6, 2, tzinfo=UTC),
        interval="daily",
        output_format="json",
    )
    db_session.add(arch)
    await db_session.commit()
    await db_session.refresh(arch)

    tmpl = _VarTemplate(
        variable_ids=json.dumps([var.id]),
        custom_start=datetime(2026, 6, 1, tzinfo=UTC),
        custom_end=datetime(2026, 6, 2, tzinfo=UTC),
    )
    await generate_report_from_template(
        tmpl, datetime(2026, 6, 1), datetime(2026, 6, 2), db_session, arch.id, lang="tr"
    )
    await db_session.refresh(arch)
    assert arch.status == "completed"
    refs = json.loads(arch.variable_refs_json)
    assert refs[0]["variable_id"] == var.id
    assert refs[0]["version"] == 7
    # compressed summary carries the variable values
    summary = json.loads(gzip.decompress(arch.result_json))
    assert summary["variables"][0]["code"] == "var_orch"
    assert summary["variables"][0]["value"] == pytest.approx(60.0)


@pytest.mark.asyncio
async def test_json_output_file_includes_variables(db_session, tmp_path, monkeypatch):
    from app.services.report_generator import generate_report_from_template

    monkeypatch.chdir(tmp_path)
    var = await _make_scalar_var(db_session)
    arch = ReportArchive(
        status="pending",
        trigger="manual",
        tag_ids="[]",
        start=datetime(2026, 6, 1, tzinfo=UTC),
        end=datetime(2026, 6, 2, tzinfo=UTC),
        interval="daily",
        output_format="json",
    )
    db_session.add(arch)
    await db_session.commit()
    await db_session.refresh(arch)
    tmpl = _VarTemplate(variable_ids=json.dumps([var.id]))
    await generate_report_from_template(
        tmpl, datetime(2026, 6, 1), datetime(2026, 6, 2), db_session, arch.id
    )
    await db_session.refresh(arch)
    with open(arch.file_path) as f:
        payload = json.load(f)
    assert payload["variables"][0]["code"] == "var_orch"


@pytest.mark.asyncio
async def test_archive_has_variable_refs_column(db_session):
    arch = ReportArchive(
        status="pending",
        trigger="manual",
        tag_ids="[]",
        start=datetime(2026, 6, 1, tzinfo=UTC),
        end=datetime(2026, 6, 2, tzinfo=UTC),
        interval="daily",
        output_format="json",
    )
    db_session.add(arch)
    await db_session.commit()
    await db_session.refresh(arch)
    assert arch.variable_refs_json is None
    arch.variable_refs_json = json.dumps([{"variable_id": 1, "code": "x", "version": 1}])
    await db_session.commit()
    await db_session.refresh(arch)
    assert json.loads(arch.variable_refs_json)[0]["code"] == "x"


@pytest.mark.asyncio
async def test_excel_output_has_variables_sheet(db_session, tmp_path, monkeypatch):
    from io import BytesIO

    from openpyxl import load_workbook

    from app.services.report_generator import generate_report_from_template

    monkeypatch.chdir(tmp_path)
    var = await _make_scalar_var(db_session)
    arch = ReportArchive(
        status="pending",
        trigger="manual",
        tag_ids="[]",
        start=datetime(2026, 6, 1, tzinfo=UTC),
        end=datetime(2026, 6, 2, tzinfo=UTC),
        interval="daily",
        output_format="excel",
    )
    db_session.add(arch)
    await db_session.commit()
    await db_session.refresh(arch)
    tmpl = _VarTemplate(output_format="excel", variable_ids=json.dumps([var.id]))
    await generate_report_from_template(
        tmpl, datetime(2026, 6, 1), datetime(2026, 6, 2), db_session, arch.id
    )
    await db_session.refresh(arch)
    with open(arch.file_path, "rb") as f:
        wb = load_workbook(BytesIO(f.read()))
    assert "Tesis Değişkenleri" in wb.sheetnames
    ws = wb["Tesis Değişkenleri"]
    codes = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "var_orch" in codes


@pytest.mark.asyncio
async def test_pdf_output_renders_with_variables(db_session, tmp_path, monkeypatch):
    """PDF çıktısında tesis değişkenleri bölümü üretilmeli (duman testi)."""
    from app.services.report_generator import generate_report_from_template

    monkeypatch.chdir(tmp_path)
    var = await _make_scalar_var(db_session)
    arch = ReportArchive(
        status="pending",
        trigger="manual",
        tag_ids="[]",
        start=datetime(2026, 6, 1, tzinfo=UTC),
        end=datetime(2026, 6, 2, tzinfo=UTC),
        interval="daily",
        output_format="pdf",
    )
    db_session.add(arch)
    await db_session.commit()
    await db_session.refresh(arch)
    tmpl = _VarTemplate(output_format="pdf", variable_ids=json.dumps([var.id]))
    await generate_report_from_template(
        tmpl, datetime(2026, 6, 1), datetime(2026, 6, 2), db_session, arch.id
    )
    await db_session.refresh(arch)
    assert arch.status == "completed"
    assert arch.file_size_bytes and arch.file_size_bytes > 0
    assert arch.file_path.endswith(".pdf")
    # WeasyPrint içerik akışlarını sıkıştırdığından (FlateDecode) ham bayt denetimi
    # yapılmıyor; bölüm varlığı test_build_pdf_variables_section_in_html ile doğrulanıyor.


def test_build_pdf_variables_section_in_html(monkeypatch):
    """build_pdf değişkenler varken HTML'e Tesis Değişkenleri bölümü eklemeli."""
    from datetime import datetime as dt
    from unittest.mock import MagicMock

    import app.services.pdf_builder as _mod

    # WeasyPrint'i monkeypatch ile yakalıyoruz; gerçek render yok
    captured: list[str] = []

    class _FakeHTML:
        def __init__(self, *, string: str):
            captured.append(string)

        def write_pdf(self) -> bytes:
            return b"%PDF-fake"

    monkeypatch.setattr(_mod, "HTML", _FakeHTML)

    archive = MagicMock()
    archive.start = dt(2026, 6, 1)
    archive.end = dt(2026, 6, 2)
    template = MagicMock()
    template.name = "Test"
    template.interval = "daily"
    template.show_summary_stats = False
    template.show_trend_charts = False
    template.show_anomaly_table = False

    variables = [
        {
            "code": "V1",
            "name": "Debi",
            "unit": "m3",
            "kind": "scalar",
            "value": 99.0,
            "points": None,
            "warning": "",
        }
    ]

    _mod.build_pdf(
        archive,
        [],
        template,
        "Test Tesisi",
        dt(2026, 6, 1),
        variables=variables,
    )

    assert captured, "HTML nesnesi oluşturulmadı"
    html = captured[0]
    # Bölüm başlığı HTML'de bulunmalı
    assert "Tesis Değişkenleri" in html
    # Değişken kodu HTML'de bulunmalı
    assert "V1" in html
