"""Report-pack generation service: section parity + non-empty PDF/Excel."""

import json
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.compliance import (
    ComplianceDischargePoint,
    ComplianceLimit,
    ComplianceParameter,
    CompliancePermit,
)
from app.models.tag import Tag, TagReading
from app.services.compliance_engine import evaluate_permit
from app.services.compliance_report import (
    SECTION_KEYS,
    build_report_pack_data,
    render_excel,
    render_json,
    render_pdf,
)


def _ts(year=2026, month=6, day=1, hour=0, minute=0) -> datetime:
    return datetime(year, month, day, hour, minute)


async def _seed_permit_with_breach(db, *, requires_explanation=False) -> int:
    permit = CompliancePermit(name="Report Permit", facility_name="WWTP-1", authority="EPA")
    point = ComplianceDischargePoint(permit=permit, code="OUT", name="Outfall")
    tag = Tag(node_id="ns=1;s=RPT_COD", name="RPT_COD")
    parameter = ComplianceParameter(
        permit=permit,
        discharge_point=point,
        parameter_name="COD",
        unit="mg/L",
        source_type="scada",
        tag=tag,
    )
    limit = ComplianceLimit(
        parameter=parameter,
        limit_type="value_limit",
        aggregation="instant",
        max_value=10.0,
        requires_explanation=requires_explanation,
    )
    db.add_all([permit, point, tag, parameter, limit])
    await db.flush()
    db.add(TagReading(tag_id=tag.id, value=12.0, quality=192, timestamp=_ts(2026, 6, 1, 12)))
    await db.flush()
    return permit.id


@pytest.mark.asyncio
async def test_json_includes_every_section_key(db_session):
    permit_id = await _seed_permit_with_breach(db_session)
    start, end = _ts(2026, 6, 1), _ts(2026, 7, 1)
    await evaluate_permit(db_session, permit_id, start, end)

    data = await build_report_pack_data(db_session, permit_id, start, end)
    for key in SECTION_KEYS:
        assert key in data, f"missing section {key}"

    payload = json.loads(render_json(data))
    for key in SECTION_KEYS:
        assert key in payload


@pytest.mark.asyncio
async def test_breach_listed_in_event_summary(db_session):
    permit_id = await _seed_permit_with_breach(db_session)
    start, end = _ts(2026, 6, 1), _ts(2026, 7, 1)
    await evaluate_permit(db_session, permit_id, start, end)

    data = await build_report_pack_data(db_session, permit_id, start, end)
    types = {e["event_type"] for e in data["event_summary"]}
    assert "limit_exceeded" in types
    assert data["period_summary"]["total_events"] >= 1


@pytest.mark.asyncio
async def test_excel_returns_nonempty_workbook(db_session):
    permit_id = await _seed_permit_with_breach(db_session)
    start, end = _ts(2026, 6, 1), _ts(2026, 7, 1)
    await evaluate_permit(db_session, permit_id, start, end)
    data = await build_report_pack_data(db_session, permit_id, start, end)

    blob = render_excel(data)
    assert isinstance(blob, bytes)
    assert len(blob) > 0
    wb = load_workbook(BytesIO(blob))
    assert "Cover" in wb.sheetnames
    assert "Events" in wb.sheetnames


@pytest.mark.asyncio
async def test_pdf_returns_nonempty_bytes(db_session):
    permit_id = await _seed_permit_with_breach(db_session)
    start, end = _ts(2026, 6, 1), _ts(2026, 7, 1)
    await evaluate_permit(db_session, permit_id, start, end)
    data = await build_report_pack_data(db_session, permit_id, start, end)

    blob = render_pdf(data)
    assert isinstance(blob, bytes)
    assert blob[:4] == b"%PDF"
