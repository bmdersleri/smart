# tests/test_excel_grafana_sheet.py
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.excel_builder import build_advanced_excel

# 1x1 PNG (geçerli) — openpyxl image yüklemesi için
_PNG_1x1 = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108020000009077"
    "53de0000000c4944415408d763f8cfc0f01f0005010101a5f6457c0000000049454e44ae426082"
)


def _tpl():
    return SimpleNamespace(
        show_summary_stats=False,
        show_raw_data=False,
    )


def test_excel_adds_grafana_sheet_with_image():
    archive = SimpleNamespace(id=1)
    out = build_advanced_excel(
        archive,
        [],
        _tpl(),
        b"",
        lang="en",
        grafana_charts=[{"title": "Debi", "png": _PNG_1x1, "error": None}],
    )
    wb = load_workbook(BytesIO(out))
    assert "Grafana" in wb.sheetnames


def test_excel_no_grafana_sheet_when_empty():
    archive = SimpleNamespace(id=1)
    out = build_advanced_excel(archive, [], _tpl(), b"", lang="en")
    wb = load_workbook(BytesIO(out))
    assert "Grafana" not in wb.sheetnames
